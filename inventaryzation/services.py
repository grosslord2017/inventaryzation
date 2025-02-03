import os
import random
import datetime as dt
from os import makedirs, listdir
from time import strftime as st
from zipapp import create_archive

from django.template.context_processors import request
from docxtpl import DocxTemplate
from django.http import Http404
from django.utils.autoreload import iter_modules_and_files
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from xlrd import open_workbook
from iris.settings import BASE_DIR, MEDIA_ROOT
from .models import INVENT_GROUP, Inventaryzation, Act, InventarNumberTemp, ArchiveUtil, Jurnal



# relations for locations
relations = {
    "accademy": "Академія (Київ)",
    "shop_riverMall": "Магазин Рівер Мол",
    "kiev_office": "Київський офіс",
    "shop_shevchenko390": "Іріс (магазин Іріс, Шевченка 390)",
    "shop_doshkevicha19": "Магазин Дашкевича 19",
    "shop_sumgaitskaya39": "Магазин Сумгаїтська 39",
    "che_basement_buh": "Черкаський офіс (каб. Сиром'ятникової)",
    "che_basement_other": "Черкаський офіс (конференц зала+коридор у підвалі)",
    "che_basement_sys": "Черкаський офіс (каб. Системного адміністратора)",
    "che_security": "Черкаський офіс (каб. СБ)",
    "che_buh": "Черкаський офіс (каб. Бухгалтерії)",
    "che_kitchen": "Черкаський офіс (кухня+кімната відпочинку)",
    "che_operators": "Черкаський офіс (каб. операторів)",
    "che_server": "Черкаський офіс (серверна)",
    "warehouse_davines": "Основний склад (Давінес, Олвейз)",
    "warehouse_electric": "Основний склад (Інструменти, Барба Італіано)",
    "warehouse_estel": "Склад митниця",
    "che_directors": "Черкаський офіс (каб. Директорів)",
}

class ParsExcel(object):
    """this class take a file in a byte code, convert him and pars.
        After that, created new items if this item is not created yet"""

    count_items = 0

    def __init__(self, file, autor):
        self.file = file
        self.autor = autor

    # def parse(self):
    #     wb = load_workbook(filename=BytesIO(self.file.read()))
    #     sheet = wb['Інвентаризація']
    #     n = 99999  # число рядков
    #     count = 0
    #     for i in range(2, n):
    #         if count >= 5:
    #             break
    #         item_atr = []
    #         for j in range(1, 12):
    #             a = sheet.cell(row=i, column=j).value
    #             item_atr.append(a)
    #         if item_atr[0] == None and item_atr[1] == None:
    #             count += 1
    #             continue
    #
    #         inv_numb = inventar_number()
    #
    #         # if item_atr[3] == '-' or None:
    #         #     item_atr[3] = inv_numb
    #
    #         try:
    #             Inventaryzation.objects.get(serial=item_atr[3])
    #             continue
    #         except:
    #             self.create_item(item_atr, inv_numb)


    # def create_item(self, item_atr, inv_numb):
    #     Inventaryzation.objects.create(
    #         group=item_atr[0] if item_atr[0] in [i[0] for i in INVENT_GROUP] else '---',
    #         name=item_atr[1],
    #         model=item_atr[2],
    #         serial=item_atr[3] if item_atr[3] != (None or '-') else inv_numb,
    #         # serial=item_atr[3],
    #         inv_numb=inv_numb,
    #         location=item_atr[4],
    #         responsible=item_atr[5],
    #         description=item_atr[6] if item_atr[6] != None else '-',
    #         price=item_atr[7] if item_atr[7] != None else 0,
    #         date_of_purchase=dt.datetime.strptime(
    #             item_atr[8], "%d.%m.%Y").strftime("%Y-%m-%d") if item_atr[8] != None else None,
    #         market_price=item_atr[9] if item_atr[9] != None else 0,
    #         date_of_registration=dt.datetime.strptime(
    #             item_atr[10], "%d.%m.%Y").strftime("%Y-%m-%d") if item_atr[10] != None else None
    #     ).save()

    # работает по началу пока не затянем все с файлов (с заранее сформироваными штрих-кодами)
    def convert_xls_to_xlsx(self, file):
        """
           Конвертирует файл .xls, представленный в виде InMemoryUploadedFile, в файл .xlsx.

           :param in_memory_file: Объект InMemoryUploadedFile (.xls файл).
           :return: BytesIO объект с содержимым .xlsx файла.
           """
        # Открываем .xls файл из памяти
        wb_xls = open_workbook(file_contents=file.read())
        wb_xlsx = Workbook()
        sheet_xlsx = wb_xlsx.active

        # Копируем данные из .xls в .xlsx
        sheet_xls = wb_xls.sheet_by_index(0)
        for row in range(sheet_xls.nrows):
            row_data = sheet_xls.row_values(row)
            sheet_xlsx.append(row_data)

        # Сохраняем результат в объект BytesIO
        xlsx_file = BytesIO()
        wb_xlsx.save(xlsx_file)
        xlsx_file.seek(0)  # Перемещаем указатель в начало файла
        return xlsx_file


    def parse_temp(self):
        format_file = str(self.file).split('.')[-1]
        if format_file == 'xls':
            xlsx_format = self.convert_xls_to_xlsx(self.file)
            wb = load_workbook(filename=BytesIO(xlsx_format.read()))
        else:
            wb = load_workbook(filename=BytesIO(self.file.read()))
        sheet = wb[wb.sheetnames[0]]
        n = 99999  # число рядков
        count = 0

        items = []

        for i in range(2, n):
            if count >= 5:
                break
            item_atr = []
            for j in range(1, 13):
                a = sheet.cell(row=i, column=j).value
                item_atr.append(a)
            if item_atr[0] == None and item_atr[1] == None:
                count += 1
                continue

            if item_atr[3] == '-' or None:
                item_atr[3] = item_atr[4]

            try:
                Inventaryzation.objects.get(serial=item_atr[3])
                continue
            except:
                try:
                    Inventaryzation.objects.get(inv_numb=item_atr[4])
                    continue
                except:
                    items.append(item_atr)
                    # self.create_item_temp(item_atr)
        self.create_item_temp(items)


    def create_item_temp(self, items):

        for item_atr in items:
            Inventaryzation.objects.create(
                group=item_atr[0] if item_atr[0] in [i[0] for i in INVENT_GROUP] else '---',
                name=item_atr[1],
                model=item_atr[2],
                serial=item_atr[3],
                inv_numb=item_atr[4],
                location=item_atr[5],
                responsible=item_atr[6],
                description=item_atr[7] if item_atr[7] != None else '-',
                price=item_atr[8] if item_atr[8] != None else 0,
                date_of_purchase=dt.datetime.strptime(
                    item_atr[9], "%d.%m.%Y").strftime("%Y-%m-%d") if item_atr[9] != None else None,
                market_price=item_atr[10] if item_atr[10] != None else 0,
                date_of_registration=dt.datetime.strptime(
                    item_atr[11], "%d.%m.%Y").strftime("%Y-%m-%d") if item_atr[11] != None else None
            ).save()
            self.count_items += 1

        create = CreateActFiles(self.autor)

#--------- Фильтрует по пользователю, разбивает на гурппы по пользователям и создает отдельный акт на пользователя со всеми его итемами------
        items_sort = sorted(items, key=lambda item_resp: item_resp[6])
        responsible = items_sort[0][6]
        items_for_one_responsible = []
        for item in items_sort:
            if item[6] != responsible:
                responsible = item[6]
                create.add_new(items_for_one_responsible)
                items_for_one_responsible.clear()
                items_for_one_responsible.append(item)
            else:
                items_for_one_responsible.append(item)
        if items_for_one_responsible:
            create.add_new(items_for_one_responsible)


        #-------------------------------------------
        # create.add_new(items)


class CreateExcel(object):
    """this class create dir 'downloads', if not excist, in a base path
     and crate excel file with content models date"""

    def __init__(self):
        self.path_to_download = f'{BASE_DIR}/downloads'
        self.workbook = Workbook()
        self.file_name = f"Таблиця_{st('%d')}-{st('%m')}-{st('%y')}_{st('%H')}-{st('%M')}-{st('%S')}"
        self.table_path = self.__create_dir()
        self.sheet = self.workbook.worksheets[0]
        self.line_count = 2
        self.header = (
            ["№п/п", "Група матеріального активу", "Назва матеріального активу", "Модель",
             "Серійний номер/VIN-номер для авто",
             "Інвентарний номер неамтеріального активу", "Місцезнаходження нематеріального активу",
             "Матеріально-відповідальна особа",
             "Опис", "Ціна придбання, з ПДВ", "Дата придбання",
             "Ринкова ціна на дату постановки активу на облік, якщо ціна придбання невідома",
             "Дата оприбуткування при інвентаризації якщо дата придбання невідома"]
        )
        self.__preset()

    def __create_dir(self):
        if "downloads" not in listdir(BASE_DIR):
            makedirs("downloads")
        self.__clear_dir()
        path = f"{self.path_to_download}/{self.file_name}.xlsx"
        return path

    def __preset(self):
        self.sheet.title = f"{st('%d')}_{st('%m')}_{st('%y')}"
        self.sheet.append(self.header)

        for i in range(1, len(self.header) + 1):
            cell_header = self.sheet.cell(row=1, column=i)
            cell_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_header.font = Font(name="Calibri", bold=True, size=9)
            self.sheet.row_dimensions[1].height = 70
            self.sheet.column_dimensions[get_column_letter(i)].width = 20

    def push_objects(self, item_atr_list):

        self.sheet.append(item_atr_list)

        for g in range(1, len(self.header) + 1):
            cell = self.sheet.cell(row=self.line_count, column=g)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=11)
            self.sheet.row_dimensions[g].height = 30
        self.line_count += 1

    def save_excel(self):
        self.workbook.save(self.table_path)

    def __clear_dir(self):
        if len(listdir(self.path_to_download)) > 0:
            for i in listdir(self.path_to_download):
                os.remove(f'{self.path_to_download}/{i}')


class CreateActFiles(object):
    transfer_template = f"{BASE_DIR}/acts_templ/transfer_template.docx"
    recycling_template = f"{BASE_DIR}/acts_templ/recycling_template.docx"
    add_new_template = f"{BASE_DIR}/acts_templ/add_new_template.docx"

    def __init__(self, autor):
        self.autor = f'{autor.username}: {autor.last_name} {autor.first_name}'

    # возможно переделать
    def chack_last_acts_id(self):
        act_id = 0
        try:
            act_id = Act.objects.all().order_by('id').last().id
        except:
            pass
        return act_id


    def transfer(self, list_objects, to_whom):
        doc = DocxTemplate(self.transfer_template)
        data = []
        from_whom = ''
        location = ''
        act_id = self.chack_last_acts_id() + 1

        for i in list_objects:
            if i.responsible != from_whom:
                from_whom = i.responsible
            if i.location != location:
                location = i.location

            temp_dict = {"name": f"{i.name} {i.model}",
                         "inv_num": f"{i.inv_numb}",
                         "price": f"{i.price if i.price != 0.0 else i.market_price}",
                         "serial_num": f"{i.serial}"}
            data.append(temp_dict)

        context = {
            "name_1": from_whom,
            "name_2": to_whom,
            "table_data": data,
            "date_day": st('%d'),
            "date_month": st('%m'),
            "date_year": st('%Y'),
            "location": location,
            "number": act_id
        }
        doc.render(context)

        path_to_file = f"acts/Акт_приймання-передачі_{act_id}.docx"
        Act.objects.create(
            operation='Зміна МВО',
            file=path_to_file
        )
        doc.save(f"{MEDIA_ROOT}/{path_to_file}")

        for item in list_objects:
            self.create_jurnal_objects(item, 'Зміна МВО', act_id, description=to_whom)


    def utilization(self, list_objects):
        print(list_objects) # [[<Inventaryzation: Inventaryzation object (9)>, '1'], [<Inventaryzation: Inventaryzation object (20)>, '6']]
        doc = DocxTemplate(self.recycling_template)
        data = []
        #-----------------------------
        location = list_objects[0][0].location
        responsible = list_objects[0][0].responsible
        #------------------------------
        act_id = self.chack_last_acts_id() + 1

        for i in range(len(list_objects)):
            temp_dict = {
                "name": f"{list_objects[i][0].name} {list_objects[i][0].model}",
                "inv_num": f"{list_objects[i][0].inv_numb}",
                "price": f"{list_objects[i][0].price if list_objects[i][0].price != 0.0 else list_objects[i][0].market_price}",
                "date_of_registration":
                    f"{list_objects[i][0].date_of_registration if list_objects[i][0].date_of_registration else list_objects[i][0].date_of_purchase}",
                "reason": f"{list_objects[i][1]}"
            }
            data.append(temp_dict)

        context = {
            "table_data": data,
            "date_day": st('%d'),
            "date_month": st('%m'),
            "date_year": st('%Y'),
            "location": location,
            "sum_of_objects": len(list_objects),
            "responsible": responsible,
            "act_id": act_id,
        }
        doc.render(context)

        path_to_file = f"acts/Акт_списання_{act_id}.docx"
        Act.objects.create(
            operation='Списання',
            file=path_to_file
        )
        doc.save(f"{MEDIA_ROOT}/{path_to_file}")

        self.create_archive_util_objects(list_objects, act_id)


    def add_new(self, lists_of_atr):
        doc = DocxTemplate(self.add_new_template)
        data = []
        workers = []
        act_id = self.chack_last_acts_id() + 1

        for list_atr in lists_of_atr:
            if list_atr[6] not in workers:
                workers.append(list_atr[6])
            temp_dict = {"name": f"{list_atr[1]} {list_atr[2]}",
                         "inv_num": f"{list_atr[4]}",
                         "price": f"{list_atr[8] if list_atr[8] != None else list_atr[10]}",
                         "serial_num": f"{list_atr[3]}",
                         "to_whom": f"{list_atr[6]}",
                         "location": f"{list_atr[5]}"}
            data.append(temp_dict)


        context = {
            "table_data": data,
            "date_day": st('%d'),
            "date_month": st('%m'),
            "date_year": st('%Y'),
            "number": act_id,
            "workers": workers
        }
        doc.render(context)

        path_to_file = f"acts/Акт_приймання_{act_id}.docx"
        Act.objects.create(
            operation='Прийняття на облік',
            file=path_to_file
        )
        doc.save(f"{MEDIA_ROOT}/{path_to_file}")

        self.create_jurnal_objects(lists_of_atr, 'Прийняття на облік', act_id)


    def create_archive_util_objects(self, items, act_id):
        for i in range(len(items)):
            ArchiveUtil.objects.create(
                group=items[i][0].group,
                name_model=items[i][0].name + ' ' + items[i][0].model,
                serial=items[i][0].serial,
                inv_numb=items[i][0].inv_numb,
                responsible=items[i][0].responsible,
                description=items[i][0].description,
                price=items[i][0].price if items[i][0].price != 0.0 else items[i][0].market_price,
                reason_for_disposal=items[i][1],
                date_of_registration=items[i][0].date_of_purchase if items[i][0].date_of_purchase else items[i][0].date_of_registration
            )

            self.create_jurnal_objects(items[i][0], 'Списання', act_id, description=items[i][1])

            items[i][0].delete()


    def create_jurnal_objects(self, item, operation, act_id, description=''):
        if operation == 'Списання':
            item_info = [f'{item.name} {item.model}', f'{item.serial}', f'{item.inv_numb}']
            Jurnal.objects.create(
                autor=self.autor,
                operation=operation,
                from_whom=item.responsible,
                to_whom='-',
                item=item_info,
                reason=description,
                act_id=Act.objects.get(id=act_id),
            ).save()
        elif operation == 'Зміна МВО':
            item_info = [f'{item.name} {item.model}', f'{item.serial}', f'{item.inv_numb}']
            Jurnal.objects.create(
                autor=self.autor,
                operation=operation,
                from_whom=item.responsible,
                to_whom=description,
                item=item_info,
                reason='-',
                act_id=Act.objects.get(id=act_id),
            ).save()
        elif operation == 'Прийняття на облік':
            for item_atr in item:
                item_info = [f'{item_atr[1]} {item_atr[2]}', f'{item_atr[3]}', f'{item_atr[4]}']
                Jurnal.objects.create(
                    autor=self.autor,
                    operation=operation,
                    from_whom='-',
                    to_whom=item_atr[6],
                    item=item_info,
                    reason='-',
                    act_id=Act.objects.get(id=act_id),
                ).save()



def inventar_number(a=10000, b=99999):
    number = random.randint(a, b) # 5 symbols
    try:
        Inventaryzation.object.get(inv_numb=number)
        inventar_number()
    except:
        try:
            InventarNumberTemp.objects.get(number=number)
            inventar_number()
        except:
            InventarNumberTemp.objects.create(number=number)
            return f'{number}'


def formating_for_ajax_response(items):
    items_list = []
    for item in items:
        items_list.append([item.id, item.group, item.name, item.model,
                           item.serial, item.inv_numb, item.location,
                           item.responsible, item.description, item.price,
                           item.date_of_purchase, item.market_price, item.date_of_registration])
    return items_list


def get_all_group():
    groups = []
    for i in INVENT_GROUP:
        groups.append(i[0])
    return groups


def get_all_locations():
    locations = []
    for i in Inventaryzation.objects.all():
        locations.append(i.location)
    locations = list(set(locations))
    locations.sort()
    return locations


def get_all_workers():
    workers = []
    for i in Inventaryzation.objects.all().order_by('responsible'):
        if i.responsible in workers:
            continue
        workers.append(i.responsible)
    workers.sort()
    return workers
